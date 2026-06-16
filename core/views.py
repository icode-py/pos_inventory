# views.py
from django.shortcuts import render
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import api_view, permission_classes, throttle_classes
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.throttling import UserRateThrottle, AnonRateThrottle, ScopedRateThrottle
import logging

logger = logging.getLogger(__name__)
from django.db.models import Sum, Count, F, FloatField, ExpressionWrapper
from django.utils.dateparse import parse_date
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
from .permissions import IsManagerOrAdmin, IsCashier, IsCashierOrManager
from .models import Category, Product, SaleTransaction, SaleItem, Staff, Restock, Customer, CustomerTransaction, LoyaltySettings, BulkDiscount, AuditLog, StoreSettings
from .serializers import (
    CategorySerializer,
    ProductSerializer,
    SaleTransactionSerializer,
    StaffSerializer,
    RestockSerializer,
    CustomerSerializer,
    CustomerTransactionSerializer,
    LoyaltySettingsSerializer,
    BulkDiscountSerializer,
    AuditLogSerializer,
    StoreSettingsSerializer,
)
from rest_framework_simplejwt.views import TokenObtainPairView
from .serializers import CustomTokenObtainPairSerializer
from django.db.models.functions import TruncDate
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


# Custom Throttle Classes
class BurstRateThrottle(UserRateThrottle):
    scope = 'burst'

class SustainedRateThrottle(UserRateThrottle):
    scope = 'sustained'


class LoginRateThrottle(AnonRateThrottle):
    scope = 'login'


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
    throttle_classes = [LoginRateThrottle]


class ProductListView(APIView):
    permission_classes = [IsAuthenticated]
    # throttle_classes = [SustainedRateThrottle]

    def get(self, request):
        products = Product.objects.all().prefetch_related('bulk_discounts')
        serializer = ProductSerializer(products, many=True)
        return Response(serializer.data)


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated, IsManagerOrAdmin]
    # throttle_classes = [SustainedRateThrottle]


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all().prefetch_related('bulk_discounts')
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    # throttle_classes = [SustainedRateThrottle]


class SaleTransactionViewSet(viewsets.ModelViewSet):
    queryset = SaleTransaction.objects.all().order_by('-created_at')
    serializer_class = SaleTransactionSerializer
    permission_classes = [IsCashierOrManager]
    pagination_class = None

    def perform_create(self, serializer):
        serializer.save(cashier=self.request.user)


class StaffViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Staff.objects.all()
    serializer_class = StaffSerializer
    permission_classes = [IsManagerOrAdmin]
    pagination_class = None


class RestockHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Restock.objects.select_related('product', 'restocked_by').order_by('-restocked_at')
    serializer_class = RestockSerializer
    permission_classes = [IsAuthenticated, IsCashierOrManager]
    pagination_class = None


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsManagerOrAdmin])
@throttle_classes([BurstRateThrottle])
def register_staff(request):
    """Create a new staff account. Requires manager or admin role."""
    try:
        serializer = StaffSerializer(data=request.data)
        if serializer.is_valid():
            user = Staff.objects.create_user(
                username=serializer.validated_data['username'],
                password=request.data.get('password'),
                is_cashier=request.data.get('is_cashier', False),
                is_manager=request.data.get('is_manager', False),
                is_admin=False,   # admins can only be created via Django admin
                is_staff=True,
                is_active=True,
            )
            logger.info('Staff account created: %s by %s', user.username, request.user.username)
            AuditLog.objects.create(
                action='CREATE', model_name='Staff',
                object_id=str(user.pk), object_repr=user.username,
                changed_by=request.user,
                changes={'role': serializer.validated_data.get('role', 'cashier')},
            )
            return Response(StaffSerializer(user).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error('register_staff error: %s', e)
        return Response({'error': 'Could not create staff account.'}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsManagerOrAdmin])
@throttle_classes([BurstRateThrottle])
def restock_product(request):
    product_id = request.data.get("product_id")
    quantity = request.data.get("quantity")

    if not product_id or not quantity:
        return Response({"error": "Missing product_id or quantity"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        product = Product.objects.get(id=product_id)
        product.stock += int(quantity)
        product.save()

        restock = Restock.objects.create(
            product=product,
            quantity_added=int(quantity),
            restocked_by=request.user
        )

        AuditLog.objects.create(
            action='CREATE',
            model_name='Restock',
            object_id=str(restock.pk),
            object_repr=f'{product.name} +{quantity}',
            changed_by=request.user,
            changes={'product': product.name, 'quantity_added': int(quantity), 'new_stock': product.stock},
        )
        logger.info('Restock: %s +%s by %s', product.name, quantity, request.user.username)

        return Response({"message": "Stock updated successfully", "stock": product.stock})
    except Product.DoesNotExist:
        return Response({"error": "Product not found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsCashierOrManager])
# @throttle_classes([SustainedRateThrottle])
def sales_report(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    cashier_id = request.GET.get("cashier_id")
    product_id = request.GET.get("product_id")

    # Validate date parameters
    if start_date:
        try:
            parse_date(start_date)
        except ValueError:
            return Response({"error": "Invalid start_date format"}, status=status.HTTP_400_BAD_REQUEST)

    if end_date:
        try:
            parse_date(end_date)
        except ValueError:
            return Response({"error": "Invalid end_date format"}, status=status.HTTP_400_BAD_REQUEST)

    sales = SaleTransaction.objects.all()

    # If no cashier_id specified and user is cashier, show only their sales
    if not cashier_id and (request.user.is_cashier and not request.user.is_manager and not request.user.is_admin):
        sales = sales.filter(cashier=request.user)

    if start_date:
        sales = sales.filter(created_at__date__gte=parse_date(start_date))
    if end_date:
        sales = sales.filter(created_at__date__lte=parse_date(end_date))
    if cashier_id:
        sales = sales.filter(cashier__id=cashier_id)

    if product_id:
        sales = sales.filter(items__product__id=product_id).distinct()

    sales_data = SaleTransactionSerializer(sales, many=True).data

    # Daily Totals
    daily_summary = (
        sales.annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(total_sales=Count("id"), total_amount=Sum("total_amount"))
        .order_by("day")
    )

    return Response({
        "sales": sales_data,
        "daily_summary": daily_summary
    })


def calculate_sales_stats(sales_queryset):
    """Helper function to calculate sales statistics without code duplication"""
    total_sales = sales_queryset.aggregate(total=Sum('total_amount'))['total'] or 0
    transaction_count = sales_queryset.count()
    average_sale = total_sales / transaction_count if transaction_count > 0 else 0
    
    return {
        'total_sales': float(total_sales),
        'transaction_count': transaction_count,
        'average_sale': float(average_sale)
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
# @throttle_classes([SustainedRateThrottle])
def store_today_sales(request):
    """Get TOTAL store sales for today - for dashboard card"""
    try:
        today = timezone.now().date()
        
        # Always get ALL sales for the store today
        sales_today = SaleTransaction.objects.filter(created_at__date=today)
        
        stats = calculate_sales_stats(sales_today)
        stats['scope'] = 'store_total'
        
        return Response(stats)
    except Exception as e:
        logger.error('store_today_sales error: %s', e)
        return Response({
            'total_sales': 0,
            'transaction_count': 0,
            'average_sale': 0,
            'scope': 'store_total',
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
# @throttle_classes([SustainedRateThrottle])
def user_today_performance(request):
    """Get INDIVIDUAL user sales for today - for layout sidebar"""
    try:
        today = timezone.now().date()
        
        # Get sales for today filtered by current user (everyone sees their own sales)
        sales_today = SaleTransaction.objects.filter(
            created_at__date=today,
            cashier=request.user
        )
        
        stats = calculate_sales_stats(sales_today)
        stats['scope'] = 'user_individual'
        
        return Response(stats)
    except Exception as e:
        logger.error('user_today_performance error: %s', e)
        return Response({
            'total_sales': 0,
            'transaction_count': 0,
            'average_sale': 0,
            'scope': 'user_individual',
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ─── Audit Mixin ────────────────────────────────────────────────────────────

class AuditMixin:
    """Mix into any ModelViewSet to auto-log create/update/delete actions."""

    def _log(self, action, instance, changes=None):
        AuditLog.objects.create(
            action=action,
            model_name=instance.__class__.__name__,
            object_id=str(instance.pk),
            object_repr=str(instance)[:200],
            changed_by=self.request.user if self.request else None,
            changes=changes or {},
        )

    def perform_create(self, serializer):
        instance = serializer.save()
        self._log('CREATE', instance)

    def perform_update(self, serializer):
        old = serializer.instance
        old_vals = {}
        for field in serializer.validated_data:
            try:
                old_vals[field] = str(getattr(old, field))
            except Exception:
                pass
        instance = serializer.save()
        changes = {}
        for field, old_val in old_vals.items():
            new_val = str(getattr(instance, field))
            if old_val != new_val:
                changes[field] = {'from': old_val, 'to': new_val}
        self._log('UPDATE', instance, changes)

    def perform_destroy(self, instance):
        self._log('DELETE', instance)
        instance.delete()


class CustomerViewSet(AuditMixin, viewsets.ModelViewSet):
    queryset = Customer.objects.all().order_by('-created_at')
    pagination_class = None  # stats cards on customers page need the full list
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated]
    # throttle_classes = [SustainedRateThrottle]
    
    def get_queryset(self):
        queryset = Customer.objects.all()
        phone = self.request.query_params.get('phone', None)
        name = self.request.query_params.get('name', None)
        
        if phone:
            queryset = queryset.filter(phone__icontains=phone)
        if name:
            queryset = queryset.filter(name__icontains=name)
            
        return queryset


class CustomerTransactionViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = CustomerTransactionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None
    
    def get_queryset(self):
        customer_id = self.request.query_params.get('customer_id')
        if customer_id:
            return CustomerTransaction.objects.filter(customer_id=customer_id).select_related('sale', 'customer').order_by('-created_at')
        return CustomerTransaction.objects.none()


class LoyaltySettingsViewSet(viewsets.ModelViewSet):
    queryset = LoyaltySettings.objects.all()
    serializer_class = LoyaltySettingsSerializer
    permission_classes = [IsAuthenticated, IsManagerOrAdmin]
    pagination_class = None


@api_view(['POST'])
@permission_classes([IsAuthenticated])
# @throttle_classes([BurstRateThrottle])
def redeem_loyalty_points(request):
    customer_id = request.data.get('customer_id')
    points_to_redeem = request.data.get('points_to_redeem')
    
    if not customer_id or not points_to_redeem:
        return Response({"error": "Missing customer_id or points_to_redeem"}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        customer = Customer.objects.get(id=customer_id)
        loyalty_settings = LoyaltySettings.objects.filter(is_active=True).first()
        
        if not loyalty_settings:
            return Response({"error": "Loyalty program is not active"}, status=status.HTTP_400_BAD_REQUEST)
        
        if customer.loyalty_points < points_to_redeem:
            return Response({"error": "Insufficient loyalty points"}, status=status.HTTP_400_BAD_REQUEST)
        
        discount_amount = (points_to_redeem * float(loyalty_settings.redemption_rate)) / 100
        
        customer.loyalty_points -= points_to_redeem
        customer.save()
        
        return Response({
            "message": "Points redeemed successfully",
            "discount_amount": discount_amount,
            "remaining_points": customer.loyalty_points
        })
        
    except Customer.DoesNotExist:
        return Response({"error": "Customer not found"}, status=status.HTTP_404_NOT_FOUND)
    

class BulkDiscountViewSet(viewsets.ModelViewSet):
    queryset = BulkDiscount.objects.all()
    serializer_class = BulkDiscountSerializer
    permission_classes = [IsAuthenticated, IsManagerOrAdmin]
    pagination_class = None


# ─── Audit-aware viewsets ────────────────────────────────────────────────────

class AuditedCategoryViewSet(AuditMixin, viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    pagination_class = None  # frontend needs full list for product forms

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsManagerOrAdmin()]


class AuditedProductViewSet(AuditMixin, viewsets.ModelViewSet):
    queryset = Product.objects.all().prefetch_related('bulk_discounts')
    serializer_class = ProductSerializer
    pagination_class = None  # POS and products page need full list for client-side search

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsManagerOrAdmin()]

    def destroy(self, request, *args, **kwargs):
        from django.db.models import ProtectedError
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
        except ProtectedError:
            return Response(
                {'error': f'"{instance.name}" has sales history and cannot be deleted.'},
                status=status.HTTP_409_CONFLICT,
            )
        return Response(status=status.HTTP_204_NO_CONTENT)


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.select_related('changed_by').all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, IsManagerOrAdmin]

    def get_queryset(self):
        qs = super().get_queryset()
        model = self.request.query_params.get('model')
        action = self.request.query_params.get('action')
        user_id = self.request.query_params.get('user_id')
        if model:
            qs = qs.filter(model_name__iexact=model)
        if action:
            qs = qs.filter(action=action.upper())
        if user_id:
            qs = qs.filter(changed_by__id=user_id)
        return qs


# ─── Low-stock alerts ────────────────────────────────────────────────────────

LOW_STOCK_THRESHOLD = 10
CRITICAL_STOCK_THRESHOLD = 5

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def low_stock_alerts(request):
    low = Product.objects.filter(stock__lte=LOW_STOCK_THRESHOLD).order_by('stock')
    data = []
    for p in low:
        data.append({
            'id': p.id,
            'name': p.name,
            'stock': p.stock,
            'severity': 'critical' if p.stock <= CRITICAL_STOCK_THRESHOLD else 'low',
            'barcode': p.barcode,
            'category': p.category.name if p.category else None,
        })
    return Response({'alerts': data, 'count': len(data)})


# ─── Bulk Excel upload ───────────────────────────────────────────────────────

TEMPLATE_HEADERS = [
    'name', 'category_name', 'price', 'cost_price', 'stock',
    'barcode', 'unit_of_measure', 'is_bulk_product',
    'bulk_quantity', 'bulk_price',
]

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsManagerOrAdmin])
def download_product_template(request):
    wb = openpyxl.Workbook()

    # ── Products sheet ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Products'

    header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Row 2 onwards is blank — users fill in their data directly
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 16)

    # ── Instructions sheet ────────────────────────────────────────────────────
    wi = wb.create_sheet(title='Instructions')
    title_font = Font(bold=True, size=13)
    bold = Font(bold=True)
    green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    red_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

    wi['A1'] = 'Product Bulk Upload — Instructions'
    wi['A1'].font = title_font

    instructions = [
        ('', ''),
        ('HOW TO USE', ''),
        ('1.', 'Go to the "Products" sheet and fill in your products starting from row 2.'),
        ('2.', 'Do NOT modify or delete row 1 (the blue header row).'),
        ('3.', 'Save the file and upload it using the "Upload Excel" button.'),
        ('', ''),
        ('COLUMN GUIDE', ''),
        ('name *', 'Required. Product name (e.g. Coca Cola 50cl)'),
        ('category_name', 'Optional. Type the category name — it will be created if it does not exist.'),
        ('price *', 'Required. Selling price per unit. Must be greater than 0.'),
        ('cost_price *', 'Required. Purchase/cost price. Must be 0 or more.'),
        ('stock *', 'Required. Current stock quantity (whole number).'),
        ('barcode *', 'Required. Unique barcode/SKU. Used to match existing products on re-upload.'),
        ('unit_of_measure', 'Optional. e.g. units, kg, litres, packs. Defaults to "units".'),
        ('is_bulk_product', 'Optional. TRUE or FALSE. Set to TRUE only for bulk-packed products.'),
        ('bulk_quantity', 'Required if bulk. Number of units in one bulk pack (e.g. 24 for a crate).'),
        ('bulk_price', 'Required if bulk. Price for the entire bulk pack.'),
        ('', ''),
        ('NOTES', ''),
        ('Re-uploading', 'If a barcode already exists, the product is UPDATED not duplicated.'),
        ('Errors', 'If any row has an error the entire upload is cancelled — nothing is saved.'),
        ('Max rows', f'Maximum {MAX_UPLOAD_ROWS} products per upload. Max file size 5 MB.'),
    ]

    for r, (label, value) in enumerate(instructions, start=2):
        wi.cell(row=r, column=1, value=label).font = bold if label and not label.startswith(('1', '2', '3')) else Font()
        wi.cell(row=r, column=2, value=value)
        if label in ('HOW TO USE', 'COLUMN GUIDE', 'NOTES'):
            wi.cell(row=r, column=1).fill = header_fill
            wi.cell(row=r, column=1).font = Font(bold=True, color='FFFFFF')
        if '*' in label:
            wi.cell(row=r, column=1).fill = green_fill
        if label == 'Errors':
            wi.cell(row=r, column=1).fill = red_fill

    wi.column_dimensions['A'].width = 22
    wi.column_dimensions['B'].width = 70

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="product_upload_template.xlsx"'
    return response


MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
MAX_UPLOAD_ROWS = 1000

@api_view(['POST'])
@permission_classes([IsAuthenticated, IsManagerOrAdmin])
def bulk_upload_products(request):
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

    if not file.name.endswith(('.xlsx', '.xls')):
        return Response({'error': 'File must be an Excel file (.xlsx or .xls)'}, status=status.HTTP_400_BAD_REQUEST)

    # Fix 2: reject oversized files before loading into memory
    if file.size > MAX_UPLOAD_BYTES:
        return Response(
            {'error': f'File too large. Maximum allowed size is 5 MB.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
    except Exception:
        return Response({'error': 'Could not read Excel file. Make sure it is a valid .xlsx file.'}, status=status.HTTP_400_BAD_REQUEST)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return Response({'error': 'File is empty or only has a header row'}, status=status.HTTP_400_BAD_REQUEST)

    header_row = [str(h).strip().lower() if h else '' for h in rows[0]]
    required = {'name', 'price', 'cost_price', 'stock', 'barcode'}
    if not required.issubset(set(header_row)):
        return Response({
            'error': f'Missing required columns. Expected: {", ".join(required)}. Got: {", ".join(h for h in header_row if h)}'
        }, status=status.HTTP_400_BAD_REQUEST)

    # Fix 3: skip header only — any row where name is blank is silently ignored,
    # no fragile hints-row detection needed
    data_rows = [(i + 2, row) for i, row in enumerate(rows[1:])
                 if str(row[header_row.index('name')] if 'name' in header_row else '').strip()]

    # Fix 2: enforce row cap after stripping blanks
    if len(data_rows) > MAX_UPLOAD_ROWS:
        return Response(
            {'error': f'Too many rows. Maximum allowed is {MAX_UPLOAD_ROWS} products per upload.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    def col(row_data, name):
        try:
            return row_data[header_row.index(name)]
        except (ValueError, IndexError):
            return None

    created, updated, errors = 0, 0, []

    # Fix 1: validate everything first, then write atomically
    validated = []
    for row_num, row in data_rows:
        name = str(col(row, 'name') or '').strip()
        try:
            price = float(col(row, 'price') or 0)
            cost_price = float(col(row, 'cost_price') or 0)
            stock = int(float(col(row, 'stock') or 0))
            barcode = str(col(row, 'barcode') or '').strip()
            category_name = str(col(row, 'category_name') or '').strip()
            unit_of_measure = str(col(row, 'unit_of_measure') or 'units').strip()
            is_bulk_raw = str(col(row, 'is_bulk_product') or 'false').strip().lower()
            is_bulk = is_bulk_raw in ('true', '1', 'yes')
            bulk_quantity = int(float(col(row, 'bulk_quantity') or 1))
            bulk_price_raw = col(row, 'bulk_price')
            bulk_price = float(bulk_price_raw) if bulk_price_raw not in (None, '', 'none') else None

            if price <= 0:
                errors.append({'row': row_num, 'name': name, 'error': 'Price must be positive'})
                continue
            if cost_price < 0:
                errors.append({'row': row_num, 'name': name, 'error': 'Cost price cannot be negative'})
                continue
            if not barcode:
                errors.append({'row': row_num, 'name': name, 'error': 'Barcode is required'})
                continue

            validated.append({
                'row_num': row_num, 'name': name, 'barcode': barcode,
                'category_name': category_name, 'price': price, 'cost_price': cost_price,
                'stock': stock, 'unit_of_measure': unit_of_measure,
                'is_bulk': is_bulk, 'bulk_quantity': bulk_quantity, 'bulk_price': bulk_price,
            })
        except Exception as e:
            errors.append({'row': row_num, 'name': name, 'error': str(e)})

    # If any row failed validation, return errors without touching the database
    if errors:
        return Response({
            'created': 0,
            'updated': 0,
            'errors': errors,
            'total_processed': 0,
            'message': 'No products were saved. Fix the errors and re-upload.',
        }, status=status.HTTP_400_BAD_REQUEST)

    # Fix 1: all rows valid — write everything in one atomic transaction
    from django.db import transaction as db_transaction
    with db_transaction.atomic():
        for item in validated:
            category = None
            if item['category_name']:
                category, _ = Category.objects.get_or_create(name=item['category_name'])

            product, was_created = Product.objects.update_or_create(
                barcode=item['barcode'],
                defaults={
                    'name': item['name'],
                    'category': category,
                    'price': item['price'],
                    'cost_price': item['cost_price'],
                    'stock': item['stock'],
                    'unit_of_measure': item['unit_of_measure'],
                    'is_bulk_product': item['is_bulk'],
                    'bulk_quantity': item['bulk_quantity'] if item['is_bulk'] else 1,
                    'bulk_price': item['bulk_price'] if item['is_bulk'] else None,
                }
            )
            AuditLog.objects.create(
                action='CREATE' if was_created else 'UPDATE',
                model_name='Product',
                object_id=str(product.pk),
                object_repr=str(product)[:200],
                changed_by=request.user,
                changes={'source': 'bulk_upload'},
            )
            if was_created:
                created += 1
            else:
                updated += 1

    return Response({
        'created': created,
        'updated': updated,
        'errors': [],
        'total_processed': created + updated,
    })


# ─── Margin / cost analytics report ─────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsManagerOrAdmin])
def margin_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    cashier_id = request.GET.get('cashier_id')

    sales = SaleTransaction.objects.all()
    if start_date:
        sales = sales.filter(created_at__date__gte=parse_date(start_date))
    if end_date:
        sales = sales.filter(created_at__date__lte=parse_date(end_date))
    if cashier_id:
        sales = sales.filter(cashier__id=cashier_id)

    sale_ids = sales.values_list('id', flat=True)
    items = SaleItem.objects.filter(transaction__in=sale_ids).select_related('product')

    # Per-product aggregation
    product_stats = {}
    for item in items:
        pid = item.product.id
        if pid not in product_stats:
            product_stats[pid] = {
                'id': pid,
                'name': item.product.name,
                'category': item.product.category.name if item.product.category else 'Uncategorized',
                'revenue': 0.0,
                'cost': 0.0,
                'units_sold': 0,
            }
        qty = item.quantity
        revenue = float(item.price_at_sale) * qty
        cost = float(item.product.cost_price) * qty
        product_stats[pid]['revenue'] += revenue
        product_stats[pid]['cost'] += cost
        product_stats[pid]['units_sold'] += qty

    for p in product_stats.values():
        p['gross_profit'] = round(p['revenue'] - p['cost'], 2)
        p['margin_pct'] = round((p['gross_profit'] / p['revenue'] * 100) if p['revenue'] else 0, 2)
        p['revenue'] = round(p['revenue'], 2)
        p['cost'] = round(p['cost'], 2)

    # Per-day aggregation
    daily_stats = {}
    for item in items:
        day = item.transaction.created_at.date().isoformat()
        if day not in daily_stats:
            daily_stats[day] = {'date': day, 'revenue': 0.0, 'cost': 0.0}
        daily_stats[day]['revenue'] += float(item.price_at_sale) * item.quantity
        daily_stats[day]['cost'] += float(item.product.cost_price) * item.quantity

    for d in daily_stats.values():
        d['gross_profit'] = round(d['revenue'] - d['cost'], 2)
        d['margin_pct'] = round((d['gross_profit'] / d['revenue'] * 100) if d['revenue'] else 0, 2)
        d['revenue'] = round(d['revenue'], 2)
        d['cost'] = round(d['cost'], 2)

    total_revenue = sum(p['revenue'] for p in product_stats.values())
    total_cost = sum(p['cost'] for p in product_stats.values())
    total_profit = total_revenue - total_cost

    return Response({
        'summary': {
            'total_revenue': round(total_revenue, 2),
            'total_cost': round(total_cost, 2),
            'gross_profit': round(total_profit, 2),
            'margin_pct': round((total_profit / total_revenue * 100) if total_revenue else 0, 2),
        },
        'by_product': sorted(product_stats.values(), key=lambda x: x['gross_profit'], reverse=True),
        'by_day': sorted(daily_stats.values(), key=lambda x: x['date']),
    })


# ─── Store Settings ───────────────────────────────────────────────────────────

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsManagerOrAdmin])
@throttle_classes([BurstRateThrottle])
def update_staff(request, pk):
    try:
        staff = Staff.objects.get(pk=pk)
    except Staff.DoesNotExist:
        return Response({'error': 'Staff not found'}, status=404)
    if staff.is_admin and not request.user.is_admin:
        return Response({'error': 'Cannot modify admin accounts'}, status=403)
    if staff.pk == request.user.pk:
        return Response({'error': 'Use profile settings to modify your own account'}, status=403)
    is_manager = request.data.get('is_manager', staff.is_manager)
    if is_manager and not staff.is_manager and not request.user.is_admin:
        return Response({'error': 'Only admins can assign manager role'}, status=403)
    staff.is_active = request.data.get('is_active', staff.is_active)
    staff.is_cashier = request.data.get('is_cashier', staff.is_cashier)
    staff.is_manager = is_manager
    staff.save()
    logger.info('Staff %s updated by %s', staff.username, request.user.username)
    AuditLog.objects.create(
        action='UPDATE', model_name='Staff',
        object_id=str(staff.pk), object_repr=staff.username,
        changed_by=request.user,
        changes={k: request.data[k] for k in ('is_cashier', 'is_manager', 'is_active') if k in request.data},
    )
    return Response(StaffSerializer(staff).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsManagerOrAdmin])
@throttle_classes([BurstRateThrottle])
def reset_staff_password(request, pk):
    try:
        staff = Staff.objects.get(pk=pk)
    except Staff.DoesNotExist:
        return Response({'error': 'Staff not found'}, status=404)
    if staff.is_admin and not request.user.is_admin:
        return Response({'error': 'Cannot reset admin passwords'}, status=403)
    if staff.pk == request.user.pk:
        return Response({'error': 'Use profile settings to change your own password'}, status=403)
    new_password = str(request.data.get('new_password', ''))
    if len(new_password) < 6:
        return Response({'error': 'Password must be at least 6 characters'}, status=400)
    staff.set_password(new_password)
    staff.save()
    logger.info('Password reset for %s by %s', staff.username, request.user.username)
    AuditLog.objects.create(
        action='UPDATE', model_name='Staff',
        object_id=str(staff.pk), object_repr=staff.username,
        changed_by=request.user,
        changes={'action': 'password_reset'},
    )
    return Response({'message': 'Password reset successfully'})


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, IsManagerOrAdmin])
@throttle_classes([BurstRateThrottle])
def delete_staff(request, pk):
    try:
        staff = Staff.objects.get(pk=pk)
    except Staff.DoesNotExist:
        return Response({'error': 'Staff not found'}, status=404)
    if staff.is_admin and not request.user.is_admin:
        return Response({'error': 'Cannot delete admin accounts'}, status=403)
    if staff.pk == request.user.pk:
        return Response({'error': 'Cannot delete your own account'}, status=403)
    username = staff.username
    AuditLog.objects.create(
        action='DELETE', model_name='Staff',
        object_id=str(staff.pk), object_repr=username,
        changed_by=request.user,
        changes={},
    )
    staff.delete()
    logger.info('Staff %s deleted by %s', username, request.user.username)
    return Response({'message': 'Staff account deleted successfully'})


# ─── Store Settings ───────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([])
def get_store_settings(request):
    return Response(StoreSettingsSerializer(StoreSettings.load()).data)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsManagerOrAdmin])
def update_store_settings(request):
    obj = StoreSettings.load()
    serializer = StoreSettingsSerializer(obj, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        AuditLog.objects.create(
            action='UPDATE', model_name='StoreSettings',
            object_id='1', object_repr=obj.name,
            changed_by=request.user,
            changes={k: request.data[k] for k in request.data},
        )
        return Response(serializer.data)
    return Response(serializer.errors, status=400)